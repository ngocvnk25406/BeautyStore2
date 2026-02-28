"""
excel_export.py - Xuất báo cáo doanh thu dạng Excel
Yêu cầu: pip install openpyxl
"""
import os
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

BASE_DIR = Path(__file__).resolve().parent.parent
EXPORT_DIR = BASE_DIR / "exports"

PINK = "FFEE609C"
PURPLE = "FFB565A7"
LIGHT_PINK = "FFFFF0F5"
WHITE = "FFFFFFFF"


def _ensure_export_dir():
    EXPORT_DIR.mkdir(exist_ok=True)


def _header_style(cell, text, bg=PINK):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        bottom=Side(style="thin", color="FFDDDDDD"),
        right=Side(style="thin", color="FFDDDDDD"),
    )


def export_revenue_excel(orders: list, products_map: dict = None,
                         customers_map: dict = None) -> str:
    """
    Xuất báo cáo doanh thu ra file Excel.
    Trả về đường dẫn file hoặc raise RuntimeError nếu thiếu openpyxl.
    """
    if not OPENPYXL_OK:
        raise RuntimeError(
            "Thư viện openpyxl chưa được cài đặt.\n"
            "Vui lòng chạy: pip install openpyxl"
        )

    _ensure_export_dir()
    filename = EXPORT_DIR / f"BaoCaoDoanhThu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb = openpyxl.Workbook()

    # ── Sheet 1: Tổng quan ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Tổng quan"

    ws1.merge_cells("A1:F1")
    title_cell = ws1["A1"]
    title_cell.value = "🧴 GLOWUP BEAUTY STORE - BÁO CÁO DOANH THU"
    title_cell.font = Font(bold=True, size=16, color=PINK[2:])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 35

    ws1["A2"].value = f"Xuất ngày: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws1["A2"].font = Font(italic=True, color="FF888888", size=9)
    ws1.merge_cells("A2:F2")

    headers = ["Mã ĐH", "Ngày", "Khách hàng", "Tạm tính", "Giảm giá", "Tổng tiền"]
    for col, h in enumerate(headers, 1):
        _header_style(ws1.cell(row=4, column=col), h)
    ws1.row_dimensions[4].height = 22

    active_orders = [o for o in orders if o.get("status") != "Đã hủy"]
    grand_total = 0
    for row_idx, o in enumerate(active_orders, 5):
        cid = o.get("customer_id", "")
        cname = ""
        if customers_map and cid in customers_map:
            cname = customers_map[cid].get("name", cid)
        subtotal = o.get("subtotal", o.get("total", 0))
        discount = o.get("discount", 0)
        total = o.get("total", 0)
        grand_total += total

        row_data = [
            o.get("order_id", ""),
            o.get("datetime", ""),
            cname or cid or "Khách lẻ",
            subtotal,
            discount,
            total,
        ]
        fill = PatternFill("solid", fgColor=LIGHT_PINK[2:] if row_idx % 2 == 0 else WHITE[2:])
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(horizontal="center" if col != 3 else "left",
                                       vertical="center")
            cell.fill = fill
            if col in (4, 5, 6):
                cell.number_format = '#,##0'

    # Hàng tổng
    total_row = len(active_orders) + 5
    ws1.cell(row=total_row, column=1, value="TỔNG CỘNG").font = Font(bold=True)
    ws1.merge_cells(f"A{total_row}:C{total_row}")
    total_cell = ws1.cell(row=total_row, column=6, value=grand_total)
    total_cell.font = Font(bold=True, color=PINK[2:], size=12)
    total_cell.number_format = '#,##0'

    # Column widths sheet 1
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 22
    ws1.column_dimensions["D"].width = 15
    ws1.column_dimensions["E"].width = 12
    ws1.column_dimensions["F"].width = 16

    # ── Sheet 2: Top sản phẩm ─────────────────────────────────────────────────
    from collections import Counter
    ws2 = wb.create_sheet("Top sản phẩm")
    ws2.merge_cells("A1:E1")
    t2 = ws2["A1"]
    t2.value = "🏆 TOP SẢN PHẨM BÁN CHẠY"
    t2.font = Font(bold=True, size=14, color=PINK[2:])
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    h2 = ["Hạng", "Tên sản phẩm", "Thương hiệu", "Đã bán (SL)", "Doanh thu"]
    for col, h in enumerate(h2, 1):
        _header_style(ws2.cell(row=3, column=col), h, PURPLE)

    counter = Counter()
    revenue_c = Counter()
    for o in active_orders:
        for item in o.get("items", []):
            pid = item.get("product_id", "")
            counter[pid] += item.get("quantity", 0)
            revenue_c[pid] += item.get("price", 0) * item.get("quantity", 0)

    for rank, (pid, qty) in enumerate(counter.most_common(15), 1):
        p = (products_map or {}).get(pid, {})
        row2 = rank + 3
        ws2.cell(row=row2, column=1, value=rank).alignment = Alignment(horizontal="center")
        ws2.cell(row=row2, column=2, value=p.get("name", pid))
        ws2.cell(row=row2, column=3, value=p.get("brand", "")).alignment = Alignment(horizontal="center")
        ws2.cell(row=row2, column=4, value=qty).alignment = Alignment(horizontal="center")
        rev_cell = ws2.cell(row=row2, column=5, value=revenue_c[pid])
        rev_cell.number_format = '#,##0'
        if rank <= 3:
            for col in range(1, 6):
                ws2.cell(row=row2, column=col).font = Font(bold=True)

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 16

    # ── Sheet 3: Doanh thu theo tháng ─────────────────────────────────────────
    ws3 = wb.create_sheet("Theo tháng")
    ws3.merge_cells("A1:C1")
    t3 = ws3["A1"]
    t3.value = "📅 DOANH THU THEO THÁNG"
    t3.font = Font(bold=True, size=14, color=PINK[2:])
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 30

    for col, h in enumerate(["Tháng", "Số đơn", "Doanh thu"], 1):
        _header_style(ws3.cell(row=3, column=col), h)

    monthly_rev = {}
    monthly_cnt = {}
    for o in active_orders:
        dt_str = o.get("datetime", "")
        try:
            parts = dt_str.split(" ")[0].split("/")
            mk = f"{parts[1]}/{parts[2]}"
        except (IndexError, ValueError):
            mk = "N/A"
        monthly_rev[mk] = monthly_rev.get(mk, 0) + o.get("total", 0)
        monthly_cnt[mk] = monthly_cnt.get(mk, 0) + 1

    for row_idx, (month, rev) in enumerate(sorted(monthly_rev.items()), 4):
        fill = PatternFill("solid", fgColor=LIGHT_PINK[2:] if row_idx % 2 == 0 else WHITE[2:])
        ws3.cell(row=row_idx, column=1, value=month).fill = fill
        ws3.cell(row=row_idx, column=2, value=monthly_cnt[month]).fill = fill
        rev_cell = ws3.cell(row=row_idx, column=3, value=rev)
        rev_cell.number_format = '#,##0'
        rev_cell.fill = fill

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 18

    wb.save(str(filename))
    return str(filename)


def export_invoice_excel(order: dict, customer: dict = None, products_map: dict = None) -> str:
    """
    Xuất hóa đơn Excel cho đơn hàng.
    Trả về đường dẫn file hoặc raise RuntimeError nếu thiếu openpyxl.
    """
    if not OPENPYXL_OK:
        raise RuntimeError(
            "Thư viện openpyxl chưa được cài đặt.\n"
            "Vui lòng chạy: pip install openpyxl"
        )

    _ensure_export_dir()
    order_id = order.get("order_id", "UNKNOWN")
    filename = EXPORT_DIR / f"HoaDon_{order_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hóa Đơn"

    thin = Side(style="thin", color="FFDDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Tiêu đề cửa hàng ──────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "GLOWUP BEAUTY STORE"
    c.font = Font(bold=True, size=18, color=PINK[2:])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:E2")
    c2 = ws["A2"]
    c2.value = "Cửa hàng Mỹ phẩm & Tư vấn Skincare"
    c2.font = Font(italic=True, size=10, color="FF888888")
    c2.alignment = Alignment(horizontal="center")

    # Dòng kẻ ngang
    for col in range(1, 6):
        cell = ws.cell(row=3, column=col)
        cell.fill = PatternFill("solid", fgColor=PINK[2:])
        cell.value = ""
    ws.row_dimensions[3].height = 4

    # ── Tiêu đề hóa đơn ───────────────────────────────────────────────────────
    ws.merge_cells("A4:E4")
    hd = ws["A4"]
    hd.value = "HÓA ĐƠN BÁN HÀNG"
    hd.font = Font(bold=True, size=14)
    hd.alignment = Alignment(horizontal="center")
    ws.row_dimensions[4].height = 28

    ws["A5"].value = "Mã đơn hàng:"
    ws["A5"].font = Font(bold=True)
    ws["B5"].value = order_id

    ws["A6"].value = "Ngày:"
    ws["A6"].font = Font(bold=True)
    ws["B6"].value = order.get("datetime", "")

    # ── Thông tin khách hàng ───────────────────────────────────────────────────
    row = 8
    ws.merge_cells(f"A{row}:E{row}")
    sec = ws[f"A{row}"]
    sec.value = "THÔNG TIN KHÁCH HÀNG"
    sec.font = Font(bold=True, color=PINK[2:])
    sec.fill = PatternFill("solid", fgColor="FFFFF0F5")
    row += 1

    if customer:
        ws[f"A{row}"].value = "Họ tên:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].value = customer.get("name", "Khách lẻ")
        row += 1
        ws[f"A{row}"].value = "SĐT:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].value = customer.get("phone", "")
        row += 1
        ws[f"A{row}"].value = "Hạng thành viên:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].value = customer.get("rank", "Đồng")
        row += 1
    else:
        ws[f"A{row}"].value = "Khách hàng:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].value = "Khách lẻ"
        row += 1

    row += 1
    # ── Bảng sản phẩm ─────────────────────────────────────────────────────────
    headers = ["STT", "Sản phẩm", "Đơn giá", "SL", "Thành tiền"]
    for col, h in enumerate(headers, 1):
        _header_style(ws.cell(row=row, column=col), h)
    ws.row_dimensions[row].height = 22
    header_row = row
    row += 1

    items = order.get("items", [])
    for i, item in enumerate(items, 1):
        pid = item.get("product_id", "")
        name = item.get("name", pid)
        if products_map and pid in products_map:
            name = products_map[pid].get("name", name)
        qty = item.get("quantity", 1)
        price = item.get("price", 0)
        subtotal_item = price * qty

        fill = PatternFill("solid", fgColor=LIGHT_PINK[2:] if i % 2 == 0 else WHITE[2:])
        row_data = [i, name, price, qty, subtotal_item]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left" if col == 2 else "center",
                vertical="center"
            )
            if col in (3, 5):
                cell.number_format = '#,##0'
        row += 1

    # ── Tổng tiền ─────────────────────────────────────────────────────────────
    row += 1
    subtotal_val = order.get("subtotal", order.get("total", 0))
    discount = order.get("discount", 0)
    discount_rate = order.get("discount_rate", 0)
    total = order.get("total", 0)

    summary = [
        ("Tạm tính:", subtotal_val),
        (f"Giảm giá ({int(discount_rate * 100)}%):", -discount),
        ("TỔNG THANH TOÁN:", total),
    ]
    for label, val in summary:
        ws.merge_cells(f"A{row}:D{row}")
        lbl_cell = ws[f"A{row}"]
        lbl_cell.value = label
        lbl_cell.alignment = Alignment(horizontal="right")
        val_cell = ws.cell(row=row, column=5, value=val)
        val_cell.number_format = '#,##0'
        val_cell.alignment = Alignment(horizontal="right")
        if label.startswith("TỔNG"):
            lbl_cell.font = Font(bold=True, size=12, color=PINK[2:])
            val_cell.font = Font(bold=True, size=12, color=PINK[2:])
            val_cell.border = Border(top=Side(style="medium", color=PINK[2:]))
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    footer = ws[f"A{row}"]
    footer.value = "Cảm ơn bạn đã mua hàng tại GLOWUP BEAUTY STORE!"
    footer.font = Font(italic=True, size=9, color="FF888888")
    footer.alignment = Alignment(horizontal="center")

    # ── Độ rộng cột ───────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 6
    ws.column_dimensions["E"].width = 16

    wb.save(str(filename))
    return str(filename)
